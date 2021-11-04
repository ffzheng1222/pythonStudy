import re
import ast
from openpyxl import load_workbook
import subprocess


ok_rule_l7_data = []
ok_rule_l4_data = []




def pars_str_l4(str_list):
    rs_info_str = ''

    for rs_index in str_list.split(","):
        rs_tmp_port, rs_tmp_ip = '', ''

        for rs_element in "".join(rs_index).split(":"):
            if '.' in rs_element:
                rs_tmp_ip = '\'rs_ip\'' + ': ' + '\'' + rs_element + '\''
            else:
                rs_tmp_port = '\'rs_port\'' + ': ' + '\'' + rs_element + '\''

        one_rs_info = '{' + rs_tmp_port + ', ' + rs_tmp_ip + '}'

        if rs_index == str_list.split(",")[-1]:
            rs_info_str += one_rs_info
        else:
            rs_info_str += one_rs_info + ', '

    ok_rs_list = rs_info_str.split("#")
    return ok_rs_list



def pars_str_l7(str_list):
    rs_info_str = ''
    str_to_list = ast.literal_eval(str_list)

    for rs_index in str_to_list:
        rs_tmp_port, rs_tmp_ip = '', ''

        for (rs_k, rs_v) in rs_index.items():
            if rs_k == "rsip":
                rs_tmp_ip = '\'rs_ip\'' + ': ' + '\'' + str(rs_v) + '\''
            elif rs_k == "rsport":
                #rs_tmp_port = '\'rs_port\'' + ': ' + '\'' + str(rs_v) + '\''
                rs_tmp_port = '\'rs_port\'' + ': ' + str(rs_v)

        #one_rs_info = '{' + rs_tmp_port + ', ' + rs_tmp_ip + '}'
        one_rs_info = '{' + rs_tmp_ip + ', ' + rs_tmp_port + '}'

        if rs_index == str_to_list[-1]:
            rs_info_str += one_rs_info
        else:
            rs_info_str += one_rs_info + ', '

    ok_rs_list = rs_info_str.split("#")
    return ok_rs_list


def format_l4_data(l4_data_list):
    # print(l4_data_list)
    # print("=======================================================")

    # rslist 待整理数据
    ori_rs_list = l4_data_list[6]
    rs_list_str_tmp = re.sub('(:[0-9]*/)', '#@#', str(ori_rs_list), re.S)
    rs_list_str = rs_list_str_tmp[:-3].replace("#@#", ",")
    rs_list = pars_str_l4(rs_list_str)

    tmp_data_list = []
    tmp_data_list.append(l4_data_list[3])
    tmp_data_list.append(l4_data_list[0])
    tmp_data_list.append(l4_data_list[1])
    tmp_data_list.append(l4_data_list[2])
    tmp_data_list.append(rs_list)

    ok_rule_l4_data.append(tmp_data_list)



def format_l7_data(l7_data_list):
    # print(l7_data_list)
    # print("=======================================================")

    # rslist 待整理数据
    ori_rs_list = l7_data_list[4]
    rs_list_str = ori_rs_list.replace("\\", ",")
    rs_list = pars_str_l7(rs_list_str)


    tmp_data_list = []
    tmp_data_list.append(l7_data_list[0])
    tmp_data_list.append(l7_data_list[1])
    tmp_data_list.append(l7_data_list[2])
    tmp_data_list.append(rs_list)

    ok_rule_l7_data.append(tmp_data_list)



def get_tgw_l7_data():
    workbook = load_workbook(r'D:\SAVE\pycharm\PycharmProjects\clb_tgw_tools\tgw_json.xlsx')
    sheets = workbook.get_sheet_names()

    if sheets[0] == "tgw_rule_l7":
        booksheet = workbook.get_sheet_by_name(u'tgw_rule_l7')
        # 获取sheet页的行数据
        rows = booksheet.rows
        tgw_rule_num = booksheet.max_row
        print("总行数为:" + str(tgw_rule_num))

        # 获取sheet页的列数据
        columns = booksheet.columns
        print("总列数为:" + str(booksheet.max_column))

        for row in rows:
            l7_data_list = []
            original_data = [col.value for col in row]
            if original_data[0] == "domain":
                continue

            if original_data[4] != 'None' and original_data[4] is not None:
                for index in original_data:
                    l7_data_list.append(index)
                format_l7_data(l7_data_list)


def get_tgw_l4_data():
    workbook = load_workbook(r'D:\SAVE\pycharm\PycharmProjects\clb_tgw_tools\tgw.xlsx')
    sheets = workbook.get_sheet_names()

    if sheets[1] == "tgw_rule":
        booksheet = workbook.get_sheet_by_name(u'tgw_rule')
        # 获取sheet页的行数据
        rows = booksheet.rows
        tgw_rule_num = booksheet.max_row
        print("总行数为:" + str(tgw_rule_num))

        # 获取sheet页的列数据
        columns = booksheet.columns
        print("总列数为:" + str(booksheet.max_column))

        for row in rows:
            l4_data_list = []
            original_data = [col.value for col in row]
            if original_data[0] == "protocol":
                continue

            if original_data[2] != 'None' and original_data[6] != 'None' and original_data[6] is not None:
                #print(original_data)
                for index in original_data:
                    l4_data_list.append(index)
                format_l4_data(l4_data_list)
