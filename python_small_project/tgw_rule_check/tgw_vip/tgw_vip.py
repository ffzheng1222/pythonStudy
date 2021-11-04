import xlwt, xlrd
import json
import re
from openpyxl import load_workbook
from xlutils.copy import copy

sheet_name_xlsx_l7 = 'tgw_vip_l7.csv'
sheet_name_xlsx_l4 = 'tgw_vip_l4.csv'


def insert_RS_data(format_rs_info):
    # 1. 准备向表中添加的数据
    format_rs_str = json.dumps(format_rs_info)
    str_tmp_1 = eval(repr(format_rs_str).replace('{"', ''))
    str_tmp_2 = eval(repr(str_tmp_1).replace('"}', ''))
    str_tmp_3 = eval(repr(str_tmp_2).replace('"', ''))
    str_tmp_4 = eval(repr(str_tmp_3).replace(':', '  : '))
    str_tmp_5 = eval(repr(str_tmp_4).replace(', ', '\\n'))
    insert_data = str(str_tmp_5)
    # print(insert_data)

    # 2. 追加write数据到sheet表格中
    workbook = xlrd.open_workbook(sheet_name_xlsx_l7)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    new_worksheet.write(rows_old, 0, insert_data)  # 追加写入数据，注意是从i+rows_old行开始写入

    # 3.保存
    new_workbook.save(sheet_name_xlsx_l7)  # 保存工作簿
    print("xls格式表格【追加】写入数据成功！")
    print("")


def go_split(str):
    # 拼接正则表达式
    str_tmp_1 = eval(repr(str).replace('\\', ''))
    str_tmp_2 = eval(repr(str_tmp_1).replace('}; ', '}#;@'))
    result = str_tmp_2.split(sep='#;@')
    # print(result)
    return result


def get_rs_ip_port(ip_port):
    ip_port_dict = {}
    ip_port_tmp = eval(repr(ip_port).replace(';', ','))
    ip_port_str = json.loads(ip_port_tmp)
    # print(type(ip_port_str))

    rsip_str = ip_port_str['rsip']
    rsport_str = str(ip_port_str['rsport'])
    ip_port_dict.setdefault(rsip_str, rsport_str)

    # print(ip_port_dict['rsip'])
    # print(ip_port_dict['rsport'])

    return ip_port_dict


def format_RS_data(rs_data, rs_ip_port_dicts):
    # 处理tgw_test_l7_sheet表格RS字段，标准显示
    rs_data_list = go_split(rs_data)
    # print(rs_data_list)
    # print("*************************************************")

    for rs_ip_data in rs_data_list:
        rs_dict = get_rs_ip_port(rs_ip_data)
        rs_ip_port_dicts.update(rs_dict)
    return rs_ip_port_dicts


def go_split_L4(str):
    # 拼接正则表达式
    result = re.sub('(:[0-9]*/)', '#@', str, re.S)
    return result


def format_RS_L4_data(rs_data, rs_ip_port_dicts):
    # 处理tgw_test_l4_sheet表格RS字段，标准显示
    rs_data_list = go_split_L4(rs_data)
    str_tmp_1 = rs_data_list[:-2]

    result = eval(repr(str_tmp_1).replace(':', '  :  '))
    return result


def handle_tgw_l7_rule(booksheet, rs_ip_port_dicts):
    # 获取sheet页的行数据
    rows = booksheet.rows
    tgw_rule_num = booksheet.max_row
    print("总行数为:" + str(booksheet.max_row))

    # 获取sheet页的列数据
    columns = booksheet.columns
    print("总列数为:" + str(booksheet.max_column))

    # 迭代所有的行
    global_workbook = xlwt.Workbook(encoding="utf-8", style_compression=0)
    global_workbook.add_sheet(sheet_name_xlsx_l7, cell_overwrite_ok=True)
    global_workbook.save(sheet_name_xlsx_l7)
    count = 0
    for row in rows:
        line = [col.value for col in row]

        if line[0] != "domain":
            count += 1
            rs_ip_port_dicts = {}
            # print("rslist --> " + line[4] + "\n")
            original_rs_data = line[4]

            if original_rs_data == 'None':
                print("count=" + str(count) + " ######################################")

            else:
                data_tmp = original_rs_data.replace("[{", "{")
                rs_data = data_tmp.replace("}]", "}")

                # 格式化tgw_test_sheet表格中的RS数据
                format_rs_info = format_RS_data(rs_data, rs_ip_port_dicts)
                print("count=" + str(count) + " ######################################")
                print(format_rs_info)

                # 插入RS格式化后的数据到新表格中
                insert_RS_data(format_rs_info)


def insert_RS_L4_data(format_rs_info):
    # 1. 准备向表中添加的数据
    format_rs_str = eval(repr(format_rs_info).replace('#@', '\\n'))
    insert_data = str(format_rs_str)
    print(insert_data)

    # 2. 追加write数据到sheet表格中
    workbook = xlrd.open_workbook(sheet_name_xlsx_l4)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    new_worksheet.write(rows_old, 0, insert_data)  # 追加写入数据，注意是从i+rows_old行开始写入

    # 3.保存
    new_workbook.save(sheet_name_xlsx_l4)  # 保存工作簿
    print("xls格式表格【追加】写入数据成功！")
    print("")


def handle_tgw_l4_rule(booksheet, rs_ip_port_dicts):
    # 获取sheet页的行数据
    rows = booksheet.rows
    tgw_rule_num = booksheet.max_row
    print("总行数为:" + str(booksheet.max_row))

    # 获取sheet页的列数据
    columns = booksheet.columns
    print("总列数为:" + str(booksheet.max_column))

    # 迭代所有的行
    global_workbook = xlwt.Workbook(encoding="utf-8", style_compression=0)
    global_workbook.add_sheet(sheet_name_xlsx_l4, cell_overwrite_ok=True)
    global_workbook.save(sheet_name_xlsx_l4)
    count = 0
    for row in rows:
        line = [col.value for col in row]

        if line[0] != "protocol":
            count += 1
            rs_ip_port_dicts = {}
            # print("rslist --> " + line[5] + "\n")
            original_rs_data = line[5]
            # print(original_rs_data)

            if original_rs_data == None:
                print("count=" + str(count) + " ######################################")

            else:
                # 格式化tgw_test_sheet表格中的RS数据
                format_rs_info = format_RS_L4_data(original_rs_data, rs_ip_port_dicts)
                print("count=" + str(count) + " ######################################")
                print(format_rs_info)

                # 插入RS格式化后的数据到新表格中
                insert_RS_L4_data(format_rs_info)


def get_RS_ip():
    # 找到需要xlsx文件的位置
    workbook = load_workbook(r'D:\SAVE\pycharm\PycharmProjects\tgw_vip\tgw.xlsx')
    sheets = workbook.get_sheet_names()  # 从名称获取sheet
    # print(sheets)

    if sheets[2] == "tgw_test_l7_sheet":
        # pass
        l7_booksheet = None
        # l7_rs_ip_port_dicts 存放表格RS字段，格式化后需要显示出来的信息
        l7_rs_ip_port_dicts = {}
        l7_booksheet = workbook.get_sheet_by_name(u'tgw_test_l7_sheet')
        handle_tgw_l7_rule(l7_booksheet, l7_rs_ip_port_dicts)

    if sheets[3] == "tgw_test_l4_sheet":
        l4_booksheet = None
        # l4_rs_ip_port_dicts 存放表格RS字段，格式化后需要显示出来的信息
        l4_rs_ip_port_dicts = {}
        l4_booksheet = workbook.get_sheet_by_name(u'tgw_test_l4_sheet')
        handle_tgw_l4_rule(l4_booksheet, l4_rs_ip_port_dicts)


def main():
    get_RS_ip()


#### main program ###
if __name__ == '__main__':
    main()
